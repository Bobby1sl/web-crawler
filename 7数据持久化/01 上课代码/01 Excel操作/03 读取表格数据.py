import openpyxl


# 根据excel文件路径读取表格
work = openpyxl.load_workbook('实例.xlsx')

# 获取工作簿对象里面有几张表格
print(work.sheetnames)

# 根据表明在工作簿中提取表格
sheet = work['Sheet']

print(sheet.max_row)  # 最大行
print(sheet.max_column)  # 最大列

# # 获取第一行数据
# for i in range(1, sheet.max_column + 1):
#     print(sheet.cell(row=1, column=i).value)

# # 获取第一列
# for j in range(1, sheet.max_row + 1):
#     print(sheet.cell(row=j, column=1).value)

for i in range(1, sheet.max_column + 1):
    for j in range(1, sheet.max_row + 1):
        print(sheet.cell(row=i, column=j).value)