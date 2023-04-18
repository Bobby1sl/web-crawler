import openpyxl  # 第三方模块  pip install openpyxl


# 1. 创建一个工作簿对象
work_book = openpyxl.Workbook()
print(work_book)

# 2. 创建表对象
sheet1 = work_book.create_sheet('表1')
print(sheet1)

"""数据填充要基于表格对象操作"""
# 方式一
# sheet1['C3'] = '哈哈'
# sheet1['A1'] = 'python'

# 方式二: 常用于循环
# sheet1.cell(row=1, column=1).value = '111111'
# sheet1.cell(row=2, column=2).value = '222222'

# 方式三: 在append()括号内穿一个序列, 那么会把这个序列作为一行写入到表格中去
sheet1.append([1,2,3,4,5])
sheet1.append((5,6,7,8,9))
sheet1.append({'A': '12345', 'B': 'abcde'})


# 3. 保存工作簿对象
work_book.save('实例.xlsx')

