import openpyxl


# 1. 创建一个工作簿对象
work_book = openpyxl.Workbook()

# 2. 创建表对象
# sheet1 = work_book.create_sheet('表1')
sheet1 = work_book.active  # 使用默认的sheet表

"""数据填充要基于表格对象操作"""
for i in range(1, 10):
    for j in range(1, i + 1):
        print(f'{j} x {i} = {j * i}', end='\t')
        sheet1.cell(row=i, column=j).value = f'{j} x {i} = {j * i}'
    print()


# 3. 保存工作簿对象
work_book.save('实例.xlsx')

